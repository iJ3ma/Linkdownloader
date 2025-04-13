import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import os
import openpyxl # Nødvendig for hyperlinks
import requests
from bs4 import BeautifulSoup # Nødvendig for website scraping
import re
from urllib.parse import urlparse, unquote, urljoin # urljoin tilføjet
import mimetypes
import traceback
import concurrent.futures
import threading # Nødvendig for thread ID
import time

# ----- STANDARD KONFIGURATION -----
DEFAULT_DOWNLOAD_TIMEOUT = 30
DEFAULT_MAX_CONCURRENT_DOWNLOADS = 10

# ----- HJÆLPEFUNKTIONER -----

def sanitize_filename(filename, is_folder=False):
    """ Fjerner eller erstatter ugyldige tegn. Lidt anderledes for mapper. """
    # Fjerner flere tegn for mapper end for filer
    if is_folder:
        # Fjerner også punktum for at undgå skjulte mapper eller problemer med extensions
        # Bevarer dog underscore og bindestreg
        sanitized = re.sub(r'[\\/*?:"<>|.,!@#$%^&()+={}\[\];\'`~]', "_", filename) # Mere aggressiv for mapper
        sanitized = re.sub(r'_+', '_', sanitized).strip('_') # Erstat multiple underscores med en enkelt
    else:
        sanitized = re.sub(r'[\\/*?:"<>|]', "_", filename)

    # Fælles begrænsninger
    max_len = 150 # Reduceret max længde for bredere kompatibilitet
    original_len = len(sanitized)
    if original_len > max_len:
        if is_folder:
             sanitized = sanitized[:max_len]
        else:
             try:
                 name, ext = os.path.splitext(sanitized)
                 # Bevar extension hvis muligt og rimelig længde
                 if len(ext) < 20 and len(ext) < original_len:
                      sanitized = name[:max_len - len(ext)] + ext
                 else: # Usandsynligt, men håndter hvis extension er for lang
                      sanitized = name[:max_len]
             except Exception: # Fallback hvis splitext fejler
                 sanitized = sanitized[:max_len]

    # Undgå tomme navne efter sanitizing
    if not sanitized:
        timestamp = str(int(time.time() * 1000))[-6:]
        sanitized = f"unnamed_{'folder' if is_folder else 'file'}_{timestamp}"

    # Undgå navne der kun består af punktummer eller starter med et
    sanitized = sanitized.lstrip('.')
    if not sanitized: # Hvis det kun var punktummer
         timestamp = str(int(time.time() * 1000))[-6:]
         sanitized = f"unnamed_{'folder' if is_folder else 'file'}_{timestamp}_dots"

    return sanitized

def get_filename_from_url(url, response):
    """ Forsøger at udlede et filnavn fra URL'en eller HTTP-headeren. """
    filename = None
    try:
        # 1. Prøv Content-Disposition headeren
        content_disposition = response.headers.get('content-disposition')
        if content_disposition:
            # Udvidet regex til at håndtere forskellige formater bedre
            fname_match = re.search(r'filename\*?=(?:(?:UTF-8|iso-8859-1)\'\')?[\'"]?([^\'"]+)[\'"]?', content_disposition, re.IGNORECASE)
            if fname_match:
                filename_from_header = unquote(fname_match.group(1), encoding='utf-8', errors='replace').strip()
                if filename_from_header:
                    # Nogle gange indeholder headeren stier, tag kun filnavnet
                    filename = os.path.basename(filename_from_header)
                    if filename:
                        # print(f"      - Filnavn fundet i Content-Disposition: {filename}")
                        return sanitize_filename(filename) # Returner straks hvis fundet

        # 2. Prøv at udlede fra URL-stien
        parsed_url = urlparse(url)
        # Afkod stien før basename for at håndtere %20 osv. korrekt
        path = unquote(parsed_url.path)
        basename = os.path.basename(path)
        # Ignorer hvis stien slutter med / eller er tom efter basename
        # Tjek også om det ligner en fil (har en extension)
        if basename and '.' in basename :
             # Fjern query parametre og fragment
             basename = basename.split('?')[0].split('#')[0]
             if basename: # Tjek igen efter split
                filename = basename
                # print(f"      - Filnavn udledt fra URL: {filename}")
                return sanitize_filename(filename) # Returner straks hvis fundet

    except Exception as e:
        print(f"      - Advarsel: Fejl under forsøg på at udlede filnavn fra header/URL: {e}")

    # 3. Fallback: Generer et navn baseret på URL og content type
    try:
        parsed_url = urlparse(url) # Sikrer at parsed_url eksisterer her
        content_type = response.headers.get('content-type', '').split(';')[0].strip()
        extension = mimetypes.guess_extension(content_type) or '.download'
        # Undgå almindelige webside-extensions i fallback navnet
        if extension.lower() in ['.html', '.htm', '.php', '.aspx', '.asp', '.cfm', '.jsp', '.do', '.action', '.shtml']:
            extension = '.download'
        # Brug domæne + en del af stien + timestamp for unikhed
        domain = parsed_url.netloc.replace('.', '_') if parsed_url.netloc else 'unknown_domain'
        path_part = sanitize_filename(os.path.basename(unquote(parsed_url.path)).split('?')[0].split('#')[0])[:30] # Tag en kort, sikker del af stien
        timestamp = str(int(time.time() * 1000))[-6:]
        # Kombiner delene, men undgå dobbelt underscore hvis path_part er tom
        filename_parts = [part for part in ["downloaded", domain, path_part, timestamp] if part]
        filename = "_".join(filename_parts) + extension
        # print(f"      - Genereret fallback-filnavn: {filename}")
        return sanitize_filename(filename) # Sanitér igen for en sikkerheds skyld
    except Exception as fallback_err:
         print(f"      - Fejl i fallback-navngenerering: {fallback_err}")
         safe_domain = re.sub(r'[^a-zA-Z0-9_-]', '_', urlparse(url).netloc) if urlparse(url).netloc else 'unknown_domain'
         timestamp = str(int(time.time() * 1000))[-6:] # Sørg for at timestamp er defineret
         return sanitize_filename(f"download_from_{safe_domain}_{timestamp}.download") # Absolut sidste udvej


def download_file_threaded(url, download_subfolder, q, timeout, source_key):
    """ Downloader fil, gemmer i download_subfolder. Returnerer resultat-tuple inkl. source_key. """
    thread_id = threading.get_ident()
    save_path = None
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(url, stream=True, timeout=timeout, allow_redirects=True, headers=headers)
        response.raise_for_status() # Tjekker for 4xx/5xx fejl

        # Tjek content type for at undgå at gemme HTML-fejlsider som filer
        content_type = response.headers.get('content-type', '').lower()
        if 'text/html' in content_type:
            try:
                preview = response.content[:512].decode('utf-8', errors='ignore')
                if '<html' in preview.lower() or '<!doctype html' in preview.lower():
                     raise ValueError(f"Modtog HTML i stedet for forventet fil (Content-Type: {content_type})")
            except Exception as html_check_err:
                 print(f"      - Advarsel under HTML check for {url}: {html_check_err}")
                 if 'text/html' in content_type: raise ValueError(f"Modtog HTML (Content-Type: {content_type})")

        filename = get_filename_from_url(url, response)
        save_path = os.path.join(download_subfolder, filename)
        counter = 1
        base_name, extension = os.path.splitext(filename)
        while os.path.exists(save_path):
            save_path = os.path.join(download_subfolder, f"{base_name}_{counter}{extension}")
            counter += 1

        # Gem filen
        with open(save_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk: f.write(chunk)

        saved_filename = os.path.basename(save_path)
        q.put(("log", f"[Thread-{thread_id}] SUCCES: Gemt {saved_filename} (fra {source_key})"))
        return (True, (url, saved_filename, source_key))

    except requests.exceptions.Timeout: reason = f"Timeout ({timeout}s)"
    except requests.exceptions.RequestException as e: reason = f"Request Fejl: {e}"
    except ValueError as e: reason = f"Værdi Fejl: {e}"
    except Exception as e: reason = f"Anden Fejl: {e}"

    # Fejl-logning
    print(f"[Thread-{thread_id}] FEJL ved download {url} (fra {source_key}): {reason}")
    q.put(("log", f"[Thread-{thread_id}] FEJL: {url} (fra {source_key}) - {reason}"))
    return (False, (url, reason, source_key))

# ----- KERNE LOGIK -----

def extract_links_from_files(excel_files_list, q):
    """ Læser hyperlinks OG tekst-URL'er fra Excel-filer. Returnerer dict {source_key: set(urls)}. """
    links_by_source_file = {}
    total_files = len(excel_files_list)
    if total_files > 0: q.put(("log", f"Starter link-ekstraktion fra {total_files} Excel fil(er)..."))

    for i, excel_path in enumerate(excel_files_list):
        base_filename = os.path.basename(excel_path)
        # source_key: Mappenavn baseret på Excel-filens navn
        source_key = sanitize_filename(f"Excel_{os.path.splitext(base_filename)[0]}", is_folder=True)

        q.put(("log", f"Læser Excel fil {i+1}/{total_files}: {base_filename} (Mappe: {source_key})"))
        links_in_this_file = set()
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # q.put(("log", f"  - Scanner ark: '{sheet_name}'")) # Undlad for mindre støj
                for row in sheet.iter_rows():
                    for cell in row:
                        original_set_size = len(links_in_this_file)
                        # Tjek Hyperlink
                        if cell.hyperlink and cell.hyperlink.target:
                            url_hl = cell.hyperlink.target
                            if isinstance(url_hl, str):
                                url_hl_cleaned = url_hl.strip()
                                if url_hl_cleaned.lower().startswith(('http://', 'https://')):
                                    links_in_this_file.add(url_hl_cleaned)
                        # Tjek Celleværdi
                        if cell.value and isinstance(cell.value, str):
                             url_val = cell.value.strip()
                             if url_val.lower().startswith(('http://', 'https://')):
                                  links_in_this_file.add(url_val)

            if links_in_this_file:
                 q.put(("log", f"  - Fundet {len(links_in_this_file)} unikke links i: {base_filename}"))
                 links_by_source_file[source_key] = links_in_this_file
            else:
                 q.put(("log", f"  - Ingen links fundet i: {base_filename}"))
        except FileNotFoundError:
             q.put(("error", f"FEJL: Filen blev ikke fundet: {base_filename}"))
        except Exception as e:
            q.put(("error", f"FEJL: Kunne ikke læse Excel '{base_filename}'. Fejl: {e}"))
            q.put(("error_detail", traceback.format_exc()))

    return links_by_source_file


def extract_links_from_website(website_url, q, timeout_seconds):
    """ Henter HTML fra URL, finder fil-lignende links. Returnerer et sæt af URLs. """
    found_links = set()
    q.put(("log", f"Scanner hjemmeside: {website_url}"))
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(website_url, timeout=timeout_seconds, headers=headers) # Bruger timeout
        response.raise_for_status()

        content_type = response.headers.get('content-type', '').lower()
        if 'html' not in content_type:
            q.put(("error", f"FEJL: URL '{website_url}' returnerede ikke HTML (Content-Type: {content_type}). Kan ikke scanne."))
            return found_links

        soup = BeautifulSoup(response.text, 'html.parser')
        links_found_on_page = 0
        web_page_extensions = ('.html', '.htm', '.php', '.aspx', '.asp', '.jsp', '.do', '.action', '.shtml', '/')

        for link_tag in soup.find_all('a', href=True):
            href = link_tag['href']
            try:
                absolute_url = urljoin(website_url, href.strip())
                parsed_link = urlparse(absolute_url)
                if parsed_link.scheme in ['http', 'https']:
                    path_basename = os.path.basename(parsed_link.path)
                    if path_basename and '.' in path_basename and not absolute_url.lower().split('?')[0].split('#')[0].endswith(web_page_extensions):
                        if absolute_url not in found_links:
                            found_links.add(absolute_url)
                            links_found_on_page += 1
            except ValueError:
                continue # Ignorer ugyldige hrefs

        q.put(("log", f"  - Fundet {links_found_on_page} nye download-lignende links på: {website_url}"))

    except requests.exceptions.Timeout:
        q.put(("error", f"FEJL: Timeout ved hentning af hjemmeside {website_url} ({timeout_seconds}s)"))
    except requests.exceptions.RequestException as e:
        q.put(("error", f"FEJL: Kunne ikke hente hjemmeside {website_url}. Fejl: {e}"))
    except Exception as e:
        q.put(("error", f"FEJL: Kunne ikke parse hjemmeside {website_url}. Fejl: {e}"))
        q.put(("error_detail", traceback.format_exc()))

    return found_links


def run_download_task(links_to_process, base_download_folder_path, q, max_workers, timeout_seconds, is_retry=False):
    """ Udfører download for links, organiseret i undermapper. """
    task_name = "Genforsøg" if is_retry else "Download"
    futures = []
    future_to_info = {}
    successful_downloads_info = []
    failed_downloads_info = []
    processed_count = 0
    total_links = 0
    try:
        links_dict = {}
        if is_retry:  # Input er [(url, reason, source_key), ...]
            for url, reason, source_key in links_to_process:
                links_dict.setdefault(source_key, set()).add(url)
            total_links = len(links_to_process)
        else:
            links_dict = links_to_process
            total_links = sum(len(urls) for urls in links_dict.values())

        if total_links == 0:
            q.put(("log", f"{task_name} afsluttet (ingen links at behandle)."))
            q.put(("results", (0, 0, [], [], is_retry)))
            q.put(("enable_buttons", True))
            return

        q.put(("progress_max", total_links))
        q.put(("progress", 0))
        q.put(("log", f"Starter {total_links} {task_name.lower()}(s) (max {max_workers} ad gangen)..."))

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            for source_key, urls_in_set in links_dict.items():
                subfolder_path = os.path.join(base_download_folder_path, source_key)
                try:
                    os.makedirs(subfolder_path, exist_ok=True)
                except OSError as e:
                    q.put(("error", f"Kunne ikke oprette mappe {subfolder_path}: {e}. Springer links fra {source_key} over."))
                    [failed_downloads_info.append((url, f"Mappe-oprettelsesfejl: {e}", source_key)) for url in urls_in_set]
                    processed_count += len(urls_in_set)
                    q.put(("progress", processed_count))
                    continue
                for url in urls_in_set:
                    future = executor.submit(download_file_threaded, url, subfolder_path, q, timeout_seconds, source_key)
                    futures.append(future)
                    future_to_info[future] = (url, source_key)

            q.put(("log", f"Alle {len(futures)} download-opgaver er sendt til trådene. Venter..."))
            for future in concurrent.futures.as_completed(futures):
                processed_count += 1
                q.put(("progress", processed_count))
                url_ctx, source_key_ctx = future_to_info[future]
                try:
                    success, detail = future.result()
                    if success:
                        successful_downloads_info.append(detail)
                    else:
                        failed_downloads_info.append(detail)
                except Exception as exc:
                    q.put(("error", f"FEJL i {task_name.lower()} tråd for {url_ctx}: {exc}"))
                    failed_downloads_info.append((url_ctx, f"Tråd Fejl: {exc}", source_key_ctx))

        # Tæl antallet af "Request Fejl" i de mislykkede downloads.
        request_error_count = sum(1 for (_, reason, _) in failed_downloads_info if "Request Fejl" in reason)
        q.put(("log", f"Antal 'Request Fejl': {request_error_count}"))

        q.put(("log", f"Alle {task_name.lower()}(s) forsøgt."))
        q.put(("results", (len(successful_downloads_info), len(failed_downloads_info), failed_downloads_info, successful_downloads_info, is_retry)))
    except Exception as e:
        q.put(("error", f"FATAL FEJL i {task_name.lower()} tråd: {e}"))
        q.put(("error_detail", traceback.format_exc()))
        fail_count = total_links
        failed_list_generic = []
        if is_retry:
            failed_list_generic = [(url_t[0], f"Processing Error: {e}", url_t[2]) for url_t in links_to_process]
        else:
            failed_list_generic = [(url, f"Processing Error: {e}", key) for key, urls in links_dict.items() for url in urls]
        q.put(("results", (0, fail_count, failed_list_generic, [], is_retry)))
    finally:
        q.put(("enable_buttons", True))


def run_processing_thread_full(excel_files_list, website_urls_list, download_folder_path, q, max_workers, timeout_seconds):
     """ Wrapper der først ekstraherer links fra filer og websites, og derefter downloader. """
     links_by_source = {}
     if excel_files_list:
          excel_links = extract_links_from_files(excel_files_list, q)
          links_by_source.update(excel_links)
     if website_urls_list:
         all_website_links = set()
         for url in website_urls_list:
              # Sender timeout_seconds med til website scanneren
              website_links = extract_links_from_website(url, q, timeout_seconds)
              all_website_links.update(website_links)
         if all_website_links:
             website_source_key = "_Website_Downloads_" # Standard mappenavn
             if len(website_urls_list) == 1: # Hvis kun én URL, brug domænet
                  try: domain = urlparse(website_urls_list[0]).netloc.replace('.', '_'); website_source_key = sanitize_filename(f"_Website_{domain}", is_folder=True) if domain else website_source_key
                  except: pass
             links_by_source.setdefault(website_source_key, set()).update(all_website_links)

     if links_by_source:
          run_download_task(links_by_source, download_folder_path, q, max_workers, timeout_seconds, is_retry=False)
     else:
          q.put(("results", (0, 0, [], [], False)))
          q.put(("log", "Færdig (ingen links fundet)."))
          q.put(("enable_buttons", True))


# ----- GUI Klassen -----

class DownloaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel & Website Link Downloader")
        self.root.geometry("800x750") # Lidt bredere for URL listbox

        self.excel_files = []
        self.website_urls = [] # Liste til website URLs
        self.download_folder = ""
        self.progress_queue = queue.Queue()
        self.processing_thread = None
        self.failed_downloads_info_last_run = []

        self.concurrency_var = tk.IntVar(value=DEFAULT_MAX_CONCURRENT_DOWNLOADS)
        self.timeout_var = tk.IntVar(value=DEFAULT_DOWNLOAD_TIMEOUT)

        style = ttk.Style()
        try: themes = style.theme_names(); style.theme_use(themes[0]) # Prøv OS standard
        except: style.theme_use('clam') # Fallback

        main_frame = ttk.Frame(root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Input Sektion (Excel + Website side om side) ---
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=5)

        # 1a. Fil Valg (Venstre)
        file_frame = ttk.LabelFrame(input_frame, text="1a. Vælg Excel Filer (.xlsx)", padding="10")
        file_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        self.select_files_button = ttk.Button(file_frame, text="Vælg Filer...", command=self.select_excel_files)
        self.select_files_button.pack(pady=2, fill=tk.X) # Fyld bredden
        self.file_list_box = tk.Listbox(file_frame, height=4, width=45) # Juster bredde
        self.file_list_box.pack(fill=tk.BOTH, expand=True, pady=(2,0))

        # 1b. Website URL Input (Højre)
        url_frame = ttk.LabelFrame(input_frame, text="1b. Tilføj Website URL", padding="10")
        url_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        self.url_entry_var = tk.StringVar() # For at kunne slette tekst nemmere
        self.url_entry = ttk.Entry(url_frame, width=40, textvariable=self.url_entry_var)
        self.url_entry.pack(pady=(2,0), ipady=2, fill=tk.X)
        self.add_url_button = ttk.Button(url_frame, text="Tilføj URL", command=self.add_website_url)
        self.add_url_button.pack(pady=2, fill=tk.X)
        self.url_list_box = tk.Listbox(url_frame, height=2, width=45) # Matcher ca. bredde
        self.url_list_box.pack(fill=tk.BOTH, expand=True, pady=(2,0))

        # 2. Mappe Valg
        folder_frame = ttk.LabelFrame(main_frame, text="2. Vælg Hoved-Download Mappe", padding="10")
        folder_frame.pack(fill=tk.X, pady=5)
        self.select_folder_button = ttk.Button(folder_frame, text="Vælg Mappe...", command=self.select_download_folder)
        self.select_folder_button.pack(side=tk.LEFT, padx=5)
        self.folder_label_var = tk.StringVar(value="Ingen mappe valgt (undermapper oprettes her)")
        self.folder_label = ttk.Label(folder_frame, textvariable=self.folder_label_var, relief=tk.SUNKEN, padding=2)
        self.folder_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # 3. Indstillinger (Sliders)
        settings_frame = ttk.LabelFrame(main_frame, text="3. Indstillinger", padding="10")
        settings_frame.pack(fill=tk.X, pady=5)
        concurrency_label = ttk.Label(settings_frame, text="Max samtidige downloads:"); concurrency_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.concurrency_scale = ttk.Scale(settings_frame, from_=1, to=30, orient=tk.HORIZONTAL, length=200, variable=self.concurrency_var, command=self.update_concurrency_label); self.concurrency_scale.grid(row=0, column=1, padx=5, pady=5, sticky=tk.EW)
        self.concurrency_value_label_var = tk.StringVar(value=f"{self.concurrency_var.get()}"); concurrency_value_label = ttk.Label(settings_frame, textvariable=self.concurrency_value_label_var, width=4); concurrency_value_label.grid(row=0, column=2, padx=5, pady=5)
        timeout_label = ttk.Label(settings_frame, text="Download Timeout (sekunder):"); timeout_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.timeout_scale = ttk.Scale(settings_frame, from_=5, to=120, orient=tk.HORIZONTAL, length=200, variable=self.timeout_var, command=self.update_timeout_label); self.timeout_scale.grid(row=1, column=1, padx=5, pady=5, sticky=tk.EW)
        self.timeout_value_label_var = tk.StringVar(value=f"{self.timeout_var.get()}s"); timeout_value_label = ttk.Label(settings_frame, textvariable=self.timeout_value_label_var, width=5); timeout_value_label.grid(row=1, column=2, padx=5, pady=5)
        settings_frame.columnconfigure(1, weight=1)

        # 4. Progress Bar
        progress_frame = ttk.Frame(main_frame, padding="5 0 5 0"); progress_frame.pack(fill=tk.X, pady=5)
        self.progress_bar_label = ttk.Label(progress_frame, text="Fremskridt:"); self.progress_bar_label.pack(side=tk.LEFT, padx=(5, 2))
        self.progress_bar = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, length=300, mode='determinate'); self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        # 5. Kontrol Knapper
        control_frame = ttk.Frame(main_frame, padding="5 0 10 0"); control_frame.pack(fill=tk.X)
        self.start_button = ttk.Button(control_frame, text="Start Download", command=self.start_initial_processing, state=tk.DISABLED); self.start_button.pack(side=tk.LEFT, padx=5)
        self.retry_button = ttk.Button(control_frame, text="Genforsøg Fejlede", command=self.start_retry_processing, state=tk.DISABLED); self.retry_button.pack(side=tk.LEFT, padx=5)

        # 6. Status Log / Resultater
        status_results_frame = ttk.LabelFrame(main_frame, text="Log og Resultater", padding="10")
        status_results_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.results_text = scrolledtext.ScrolledText(status_results_frame, height=15, width=80, wrap=tk.WORD, state=tk.DISABLED); self.results_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_to_results("Klar. Vælg filer og/eller tilføj URLs, vælg mappe og juster indstillinger.")

        self.check_queue()

    def update_concurrency_label(self, value): self.concurrency_value_label_var.set(f"{int(float(value))}")
    def update_timeout_label(self, value): self.timeout_value_label_var.set(f"{int(float(value))}s")

    def select_excel_files(self):
        selected = filedialog.askopenfilenames(
            title="Vælg Excel fil(er)",
            filetypes=[("Excel filer", "*.xlsx"), ("Alle filer", "*.*")]
        )
        if selected:
            self.excel_files = list(selected)
            self.file_list_box.delete(0, tk.END)
            # Vis kun filnavne, ikke fulde stier
            display_files = [os.path.basename(f) for f in self.excel_files]
            max_display = 50
            if len(display_files) > max_display:
                for f_disp in display_files[:max_display]:
                    self.file_list_box.insert(tk.END, f_disp)
                self.file_list_box.insert(tk.END, f"... ({len(display_files) - max_display} flere)")
            else:
                for f_disp in display_files:
                    self.file_list_box.insert(tk.END, f_disp)

            self.update_start_button_state()
            self.retry_button.config(state=tk.DISABLED)
            self.failed_downloads_info_last_run = []
            self.clear_log_and_results()

    def add_website_url(self):
        url = self.url_entry_var.get().strip() # Brug var til at få værdi
        if url.lower().startswith(('http://', 'https://')):
            if url not in self.website_urls:
                 self.website_urls.append(url); self.url_list_box.insert(tk.END, url); self.url_entry_var.set("") # Ryd inputfelt via var
                 self.update_start_button_state(); self.retry_button.config(state=tk.DISABLED); self.failed_downloads_info_last_run = []; self.clear_log_and_results()
            else: messagebox.showinfo("URL Eksisterer", "Denne URL er allerede tilføjet.")
        else: messagebox.showwarning("Ugyldig URL", "Indtast venligst en gyldig URL (starter med http:// eller https://).")

    def select_download_folder(self):
        selected = filedialog.askdirectory(title="Vælg HOVED mappe til downloads")
        if selected: self.download_folder = selected; self.folder_label_var.set(selected); self.update_start_button_state(); self.retry_button.config(state=tk.DISABLED); self.failed_downloads_info_last_run = []; self.clear_log_and_results()

    def update_start_button_state(self): self.start_button.config(state=tk.NORMAL) if (self.excel_files or self.website_urls) and self.download_folder else self.start_button.config(state=tk.DISABLED)

    def clear_log_and_results(self):
        self.results_text.config(state=tk.NORMAL); self.results_text.delete('1.0', tk.END); self.log_to_results("Klar."); self.results_text.config(state=tk.DISABLED); self.progress_bar['value'] = 0

    def disable_controls(self):
        for btn in [self.select_files_button, self.select_folder_button, self.add_url_button, self.start_button, self.retry_button]: btn.config(state=tk.DISABLED)
        for scale in [self.concurrency_scale, self.timeout_scale]: scale.config(state=tk.DISABLED)
        self.url_entry.config(state=tk.DISABLED)

    def enable_controls(self):
        for btn in [self.select_files_button, self.select_folder_button, self.add_url_button]: btn.config(state=tk.NORMAL)
        for scale in [self.concurrency_scale, self.timeout_scale]: scale.config(state=tk.NORMAL)
        self.url_entry.config(state=tk.NORMAL)
        self.retry_button.config(state=tk.NORMAL) if self.failed_downloads_info_last_run else self.retry_button.config(state=tk.DISABLED)
        self.update_start_button_state() # Start knap styres af om der er input

    def start_initial_processing(self):
        if not self.excel_files and not self.website_urls: messagebox.showwarning("Input Mangler", "Vælg venligst mindst én Excel fil eller tilføj en Website URL."); return
        if not self.download_folder: messagebox.showwarning("Input Mangler", "Vælg venligst en download mappe."); return
        self.disable_controls(); self.clear_log_and_results(); self.failed_downloads_info_last_run = []; max_workers = self.concurrency_var.get(); timeout = self.timeout_var.get(); self.log_to_results(f"Starter kørsel (Max tråde: {max_workers}, Timeout: {timeout}s)...")
        # Klon listerne for at undgå race conditions hvis brugeren ændrer dem mens tråden kører
        excel_files_copy = list(self.excel_files)
        website_urls_copy = list(self.website_urls)
        self.processing_thread = threading.Thread(
            target=run_processing_thread_full,
            args=(excel_files_copy, website_urls_copy, self.download_folder, self.progress_queue, max_workers, timeout),
            daemon=True
        )
        self.processing_thread.start()

    def start_retry_processing(self):
        if not self.failed_downloads_info_last_run: messagebox.showinfo("Ingen Fejl", "Der er ingen fejlede downloads at genprøve."); return
        if not self.download_folder: messagebox.showwarning("Mappe Mangler", "Vælg venligst en download mappe."); return
        self.disable_controls(); self.log_to_results("\n--- STARTER GENFORSØG AF FEJLEDE ---")
        self.failed_downloads_info_last_run = [];  # Nulstil listen
        max_workers = self.concurrency_var.get(); timeout = self.timeout_var.get();
        # Input til run_download_task er listen af tuples: [(url, reason, source_key), ...]
        failed_to_retry = list(self.failed_downloads_info_last_run) # Kopiér listen
        self.failed_downloads_info_last_run = []; # Nulstil listen
        max_workers = self.concurrency_var.get(); timeout = self.timeout_var.get(); self.log_to_results(f"Genforsøger {len(failed_to_retry)} links (Max tråde: {max_workers}, Timeout: {timeout}s)...")
        self.processing_thread = threading.Thread(target=run_download_task, args=(failed_to_retry, self.download_folder, self.progress_queue, max_workers, timeout, True), daemon=True); self.processing_thread.start()

    def check_queue(self):
        try:
            while True:
                message_type, data = self.progress_queue.get_nowait()
                if message_type == "log":
                    self.log_to_results(data)
                elif message_type == "progress":
                    self.progress_bar['value'] = data
                    self.root.update_idletasks()
                elif message_type == "progress_max":
                    max_val = data if data > 0 else 1
                    self.progress_bar['maximum'] = max_val
                    self.progress_bar['value'] = 0
                    self.root.update_idletasks()
                elif message_type == "results":
                    success_count, fail_count, failed_info, successful_info, is_retry = data
                    self.display_results(success_count, fail_count, failed_info, successful_info, is_retry)
                    self.failed_downloads_info_last_run = failed_info
                elif message_type == "error":
                    self.log_error_to_results(f"FEJL: {data}")
                elif message_type == "enable_buttons":
                    self.enable_controls()
        except queue.Empty:
            pass
        finally:
            self.root.after(100, self.check_queue)

    def log_to_results(self, message):
        try:
            self.results_text.config(state=tk.NORMAL)
            self.results_text.insert(tk.END, f"{message}\n")
            self.results_text.config(state=tk.DISABLED)
            self.results_text.see(tk.END)
        except tk.TclError as e:
            print(f"GUI TclError: {e}")

    def log_error_to_results(self, error_message):
        self.log_to_results(f"*** {error_message} ***")

    def display_results(self, success_count, fail_count, failed_info, successful_info, is_retry):
        self.results_text.config(state=tk.NORMAL)
        result_header = "RESULTATER (GENFORSØG)" if is_retry else "RESULTATER"
        self.results_text.insert(tk.END, f"\n--- {result_header} ---\n")
        self.results_text.insert(tk.END, f"Succesfulde: {success_count}\n")
        self.results_text.insert(tk.END, f"Mislykkede: {fail_count}\n")
        self.results_text.insert(tk.END, "------------------\n")
        if failed_info:
            self.results_text.insert(tk.END, "\nMISLYKKEDE DOWNLOADS:\n")
            for url, reason, source_key in failed_info:
                self.results_text.insert(tk.END, f"- Kilde: {source_key}\n  URL: {url}\n  Årsag: {reason}\n")
        if successful_info:
            self.results_text.insert(tk.END, "\nSUCCESFULDE DOWNLOADS (Filnavn (Kilde)):\n")
            for url, filename, source_key in successful_info:
                self.results_text.insert(tk.END, f"- {filename} (Fra: {source_key})\n")
        self.results_text.insert(tk.END, "\n------------------\n")
        self.results_text.config(state=tk.DISABLED)
        self.results_text.see(tk.END)


# ----- Start GUI Applikationen -----
if __name__ == "__main__":
    root = tk.Tk()
    app = DownloaderApp(root)
    root.mainloop()